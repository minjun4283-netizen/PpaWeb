#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


def is_blank(v) -> bool:
    return v is None or str(v).strip() == ""


def normalize_name(text: str) -> str:
    return str(text or "").replace(" ", "").replace("_", "").replace("-", "").strip().lower()


def find_sheet_name(wb, target_name: str) -> str | None:
    if target_name in wb.sheetnames:
        return target_name

    target_norm = normalize_name(target_name)
    target_no_t = normalize_name(target_name.replace("T_", ""))

    for s in wb.sheetnames:
        ns = normalize_name(s)
        if ns == target_norm or ns == target_no_t:
            return s

    for s in wb.sheetnames:
        ns = normalize_name(s)
        if target_no_t and (target_no_t in ns or ns in target_no_t):
            return s

    return None


def detect_header_row_and_headers(ws) -> Tuple[int, List[Tuple[int, str]]]:
    max_scan_rows = min(ws.max_row or 1, 20)
    max_col = ws.max_column or 1

    best_row = 1
    best_headers: List[Tuple[int, str]] = []

    for row_idx in range(1, max_scan_rows + 1):
        row_headers: List[Tuple[int, str]] = []
        seen = set()

        for col_idx in range(1, max_col + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if is_blank(v):
                continue

            name = str(v).strip()
            if not name or name in seen:
                continue

            seen.add(name)
            row_headers.append((col_idx, name))

        if len(row_headers) > len(best_headers):
            best_row = row_idx
            best_headers = row_headers

    return best_row, best_headers


def extract_rows(ws) -> Tuple[List[str], List[Dict[str, str]]]:
    header_row, header_pairs = detect_header_row_and_headers(ws)
    headers = [name for _, name in header_pairs]

    rows: List[Dict[str, str]] = []
    max_row = ws.max_row or header_row

    for row_idx in range(header_row + 1, max_row + 1):
        obj: Dict[str, str] = {}
        non_empty = 0

        for col_idx, name in header_pairs:
            v = ws.cell(row=row_idx, column=col_idx).value
            text = "" if is_blank(v) else str(v).strip()
            if text:
                non_empty += 1
            obj[name] = text

        if non_empty > 0:
            rows.append(obj)

    return headers, rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsm", required=True)
    parser.add_argument("--table", required=True)
    args = parser.parse_args()

    try:
        wb = load_workbook(args.xlsm, read_only=True, data_only=True, keep_vba=True)
        try:s
            sheet_name = find_sheet_name(wb, args.table)
            if not sheet_name:
                print(json.dumps({"ok": False, "error": f"Worksheet not found: {args.table}"}, ensure_ascii=False))
                sys.exit(1)

            ws = wb[sheet_name]
            headers, rows = extract_rows(ws)
            print(json.dumps({"ok": True, "headers": headers, "rows": rows}, ensure_ascii=False))
        finally:
            wb.close()

    except Exception as e:
        print(json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()