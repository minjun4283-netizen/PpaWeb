#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ppa_loader.py — xlsm 또는 CSV 폴더에서 6개 표 데이터를 읽어옵니다.

두 가지 입력 방식을 지원합니다:
  - xlsm: openpyxl 필요. 이 폴더의 vendor/ 안에 openpyxl을 이미 동봉해뒀으므로
          pip install이나 인터넷 연결 없이도 그대로 동작합니다(순수 Python
          패키지라 별도 설치 없이 폴더째로 옮겨도 됩니다). 시트 이름
          (T_발전소 등)과 헤더 텍스트로 자동 매칭하고, 검증 열(PK중복 등)은
          자동으로 무시합니다.
  - CSV 폴더: 표준 라이브러리만 사용 (vendor/도 필요 없음). 폴더 안에
          T_발전소.csv, T_구매계약.csv ... 형식으로 파일이 있어야 함
          (헤더는 실제 컬럼명과 정확히 일치).

인터넷이 아예 안 되는 환경이라도, 이 폴더(vendor/ 포함)를 통째로 옮기면
xlsm 방식도 pip 없이 그대로 동작합니다. Python 인터프리터 자체가 없다면
python-3.x-embed-amd64.zip(공식 배포, https://www.python.org/downloads/windows/
"Windows embeddable package") 을 이 폴더 옆에 같이 풀어서 동봉하세요.
"""
import csv
import datetime
import os
import sys

# 동봉된 openpyxl/et_xmlfile을 시스템에 설치된 버전보다 우선 사용 —
# pip install 없이도, 이 폴더를 통째로 옮기기만 하면 xlsm 방식이 동작합니다.
_VENDOR_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor")
if os.path.isdir(_VENDOR_DIR) and _VENDOR_DIR not in sys.path:
    sys.path.insert(0, _VENDOR_DIR)

from ppa_schema import TABLES


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def load_from_xlsm(path: str) -> dict[str, list[dict]]:
    import openpyxl  # optional dependency, only needed for this path

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    tables_data: dict[str, list[dict]] = {}

    for t in TABLES:
        if t.key not in wb.sheetnames:
            print(f"[건너뜀] 시트를 찾을 수 없음: {t.key}")
            tables_data[t.key] = []
            continue

        ws = wb[t.key]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            tables_data[t.key] = []
            continue

        header = [str(h).strip() if h is not None else "" for h in header]
        col_index = {name: idx for idx, name in enumerate(header)}
        unmatched = [h for h in header if h and h not in t.columns]

        rows: list[dict] = []
        for raw in rows_iter:
            record: dict[str, str] = {}
            has_any = False
            for col in t.columns:
                idx = col_index.get(col)
                val = raw[idx] if idx is not None and idx < len(raw) else None
                val = _normalize_cell(val)
                record[col] = val
                if val != "":
                    has_any = True
            if has_any:
                rows.append(record)

        tables_data[t.key] = rows
        note = f" 인식 안 된 헤더(검증열 등, 정상): {', '.join(unmatched)}" if unmatched else ""
        print(f"{t.key}: {len(rows)}행 인식.{note}")

    return tables_data


def load_from_csv_dir(dir_path: str) -> dict[str, list[dict]]:
    tables_data: dict[str, list[dict]] = {}

    for t in TABLES:
        path = os.path.join(dir_path, f"{t.key}.csv")
        if not os.path.exists(path):
            print(f"[건너뜀] 파일을 찾을 수 없음: {path}")
            tables_data[t.key] = []
            continue

        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            unmatched = [h for h in (reader.fieldnames or []) if h and h not in t.columns]
            rows = []
            for raw in reader:
                record = {col: (raw.get(col) or "").strip() for col in t.columns}
                if any(v for v in record.values()):
                    rows.append(record)

        tables_data[t.key] = rows
        note = f" 인식 안 된 헤더(검증열 등, 정상): {', '.join(unmatched)}" if unmatched else ""
        print(f"{t.key}: {len(rows)}행 인식.{note}")

    return tables_data
