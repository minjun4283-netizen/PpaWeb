#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


def is_blank(v) -> bool:
    return v is None or str(v).strip() == ""


def normalize_name(text: str) -> str:
    return str(text or "").replace(" ", "").replace("_", "").replace("-", "").strip().lower()


def find_sheet_name(wb, target_name: str) -> str | None:
    if target_name in wb.sheetnames:
        return target_name

    target_norm = normalize_name(target_name)
    target_no_t = normalize_name(target_name.replace("T_", ""))

    for s in wb.sheetnames:
        ns = normalize_name(s)
        if ns == target_norm or ns == target_no_t:
            return s

    for s in wb.sheetnames:
        ns = normalize_name(s)
        if target_no_t and (target_no_t in ns or ns in target_no_t):
            return s

    return None


def detect_header_row_and_headers(ws) -> Tuple[int, List[Tuple[int, str]]]:
    max_scan_rows = min(ws.max_row or 1, 20)
    max_col = ws.max_column or 1

    best_row = 1
    best_headers: List[Tuple[int, str]] = []

    for row_idx in range(1, max_scan_rows + 1):
        row_headers: List[Tuple[int, str]] = []
        seen = set()

        for col_idx in range(1, max_col + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if is_blank(v):
                continue

            name = str(v).strip()
            if not name or name in seen:
                continue

            seen.add(name)
            row_headers.append((col_idx, name))

        if len(row_headers) > len(best_headers):
            best_row = row_idx
            best_headers = row_headers

    return best_row, best_headers


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsm", required=True)
    parser.add_argument("--table", required=True)
    parser.add_argument("--pk", required=True)
    parser.add_argument("--payload-file", required=True)
    args = parser.parse_args()

    temp_dir = None
    temp_xlsm = None

    try:
        payload = json.loads(Path(args.payload_file).read_text(encoding="utf-8"))
        pk_name = args.pk
        pk_value = str(payload.get(pk_name) or "").strip()

        if not pk_value:
            raise ValueError(f"{pk_name}는 필수입니다.")

        temp_dir = tempfile.mkdtemp(prefix="ppa_xlsm_")
        temp_xlsm = os.path.join(temp_dir, "source.xlsm")
        shutil.copy2(args.xlsm, temp_xlsm)

        wb = load_workbook(temp_xlsm, keep_vba=True, data_only=False)
        try:
            sheet_name = find_sheet_name(wb, args.table)
            if not sheet_name:
                raise ValueError(f"Worksheet not found: {args.table}")

            ws = wb[sheet_name]
            header_row, header_pairs = detect_header_row_and_headers(ws)
            header_map = {name: col_idx for col_idx, name in header_pairs}

            if pk_name not in header_map:
                raise ValueError(f"PK column not found: {pk_name}")

            pk_col = header_map[pk_name]
            target_row = None

            for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
                existing = ws.cell(row=row_idx, column=pk_col).value
                if str(existing or "").strip() == pk_value:
                    target_row = row_idx
                    break

            action = "updated"
            if target_row is None:
                target_row = (ws.max_row or header_row) + 1
                if target_row <= header_row:
                    target_row = header_row + 1
                action = "inserted"

            for key, value in payload.items():
                if key in header_map:
                    ws.cell(row=target_row, column=header_map[key]).value = "" if value is None else str(value)

            wb.save(temp_xlsm)
        finally:
            wb.close()

        shutil.copy2(temp_xlsm, args.xlsm)

        print(json.dumps({
            "ok": True,
            "result": {
                "action": action,
                "row": target_row,
                "pk_name": pk_name,
                "pk_value": pk_value,
                "table": args.table,
            }
        }, ensure_ascii=False))

    except Exception as e:
        print(json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False))
        sys.exit(1)

    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    main()