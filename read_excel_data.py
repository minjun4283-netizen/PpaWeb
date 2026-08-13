#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import sys
from typing import Dict, List, Tuple

from openpyxl import load_workbook


TARGET_TABLES = [
    "T_발전소",
    "T_구매계약",
    "T_수요기업",
    "T_판매계약",
    "T_전기사용지",
    "T_수급매칭",
]


def is_blank(v) -> bool:
    return v is None or str(v).strip() == ""


def normalize_name(text: str) -> str:
    return (
        str(text or "")
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
        .strip()
        .lower()
    )


def find_sheet_name(wb, target_name: str) -> str | None:
    if target_name in wb.sheetnames:
        return target_name

    target_norm = normalize_name(target_name)
    target_no_t = normalize_name(target_name.replace("T_", ""))

    for s in wb.sheetnames:
        ns = normalize_name(s)
        if ns == target_norm or ns == target_no_t:
            return s

    for s in wb.sheetnames:
        ns = normalize_name(s)
        if target_no_t and (target_no_t in ns or ns in target_no_t):
            return s

    return None


def detect_header_row_and_headers(ws) -> Tuple[int, List[Tuple[int, str]]]:
    max_scan_rows = min(ws.max_row or 1, 20)
    max_col = ws.max_column or 1

    best_row = 1
    best_headers: List[Tuple[int, str]] = []

    for row_idx in range(1, max_scan_rows + 1):
        row_headers: List[Tuple[int, str]] = []
        seen = set()

        for col_idx in range(1, max_col + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if is_blank(v):
                continue

            name = str(v).strip()
            if not name:
                continue
            if name in seen:
                continue

            seen.add(name)
            row_headers.append((col_idx, name))

        if len(row_headers) > len(best_headers):
            best_row = row_idx
            best_headers = row_headers

    return best_row, best_headers


def extract_rows(ws) -> Tuple[List[str], List[Dict[str, str]]]:
    header_row, header_pairs = detect_header_row_and_headers(ws)
    headers = [name for _, name in header_pairs]

    rows: List[Dict[str, str]] = []
    max_row = ws.max_row or header_row

    for row_idx in range(header_row + 1, max_row + 1):
        obj: Dict[str, str] = {}
        non_empty = 0

        for col_idx, name in header_pairs:
            v = ws.cell(row=row_idx, column=col_idx).value
            text = "" if is_blank(v) else str(v).strip()
            if text:
                non_empty += 1
            obj[name] = text

        if non_empty > 0:
            rows.append(obj)

    return headers, rows


def load_table_rows(xlsm_path: str, table_name: str) -> Tuple[List[str], List[Dict[str, str]]]:
    wb = load_workbook(
        xlsm_path,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        sheet_name = find_sheet_name(wb, table_name)
        if not sheet_name:
            return [], []

        ws = wb[sheet_name]
        return extract_rows(ws)
    finally:
        wb.close()


def cmd_lookup(xlsm_path: str):
    _, purchase_rows = load_table_rows(xlsm_path, "T_구매계약")
    _, site_rows = load_table_rows(xlsm_path, "T_전기사용지")
    _, matching_rows = load_table_rows(xlsm_path, "T_수급매칭")

    purchase_options = []
    seen_purchase = set()
    for r in purchase_rows:
        value = str(r.get("구매계약ID") or "").strip()
        if not value or value in seen_purchase:
            continue
        seen_purchase.add(value)

        extra = str(r.get("발전소ID") or "").strip()
        label = f"{value} | {extra}" if extra else value
        purchase_options.append({"value": value, "label": label})

    site_options = []
    seen_site = set()
    for r in site_rows:
        value = str(r.get("전기사용지ID") or "").strip()
        if not value or value in seen_site:
            continue
        seen_site.add(value)

        extra = str(r.get("전기사용지명") or "").strip()
        label = f"{value} | {extra}" if extra else value
        site_options.append({"value": value, "label": label})

    matching_records = []
    seen_matching = set()
    for r in matching_rows:
        value = str(r.get("수급매칭ID") or "").strip()
        if not value or value in seen_matching:
            continue
        seen_matching.add(value)

        extra = str(r.get("현황") or "").strip()
        label = f"{value} | {extra}" if extra else value
        matching_records.append({"value": value, "label": label})

    print(json.dumps({
        "ok": True,
        "purchaseOptions": purchase_options,
        "siteOptions": site_options,
        "matchingRecords": matching_records,
    }, ensure_ascii=False))


def cmd_get(xlsm_path: str, table_name: str, pk_name: str, pk_value: str):
    _, rows = load_table_rows(xlsm_path, table_name)
    target = str(pk_value or "").strip()

    for r in rows:
        value = str(r.get(pk_name) or "").strip()
        if value == target:
            print(json.dumps({"ok": True, "record": r}, ensure_ascii=False))
            return

    print(json.dumps({"ok": True, "record": {}}, ensure_ascii=False))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsm", required=True)
    parser.add_argument("--mode", required=True, choices=["lookup", "get"])
    parser.add_argument("--table", default="")
    parser.add_argument("--pk-name", default="")
    parser.add_argument("--pk-value", default="")
    args = parser.parse_args()

    try:
        if args.mode == "lookup":
            cmd_lookup(args.xlsm)
        elif args.mode == "get":
            if not args.table or not args.pk_name:
                raise ValueError("--table, --pk-name are required for mode=get")
            cmd_get(args.xlsm, args.table, args.pk_name, args.pk_value)
        else:
            raise ValueError("invalid mode")
    except Exception as e:
        print(json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()